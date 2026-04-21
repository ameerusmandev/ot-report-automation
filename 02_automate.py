import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import os
import time
from timing_logger import TimeLogger

# Create timestamp folder at start
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
output_folder = os.path.join("output_data", timestamp)
os.makedirs(output_folder, exist_ok=True)

# Initialize time logger
timer = TimeLogger(log_file=os.path.join(output_folder, "timing_log.csv"))
timer.start()

file_path = 'input_data/Source File 4-13.xlsx'

sheet_names = [
    "EDFT",
    "TH Summary",
    "TH Hourly",
    "ML Store",
    "Emp List",
    "Apr Unapr OT",
    "Check Pivot",
    "Current Week OT",
    "Sheet10",
    "SCH",
    "SCH for MO"
]

dfs = {}

read_start = time.time()
for sheet in sheet_names:
    dfs[sheet] = pd.read_excel(file_path, sheet_name=sheet)
timer.log_task("Read Excel sheets", time.time() - read_start)


dfs.keys()

# Access example:
df_edft = dfs["EDFT"]
df_th_summary = dfs["TH Summary"]
df_th_hourly = dfs["TH Hourly"]

# df_edft.head(10)


# ----------------------------
# PREP DATA
# ----------------------------
prep_start = time.time()
df = df_th_hourly.copy()

df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
df["Hours"] = pd.to_numeric(df["Hours"], errors="coerce")
timer.log_task("Prepare data", time.time() - prep_start)

# ----------------------------
# PIVOT (EE Code + Name)
# ----------------------------
pivot_start = time.time()
pivot_df = df.pivot_table(
    index=["EE Code", "Employee_Name"],
    columns="Date",
    values="Hours",
    aggfunc="sum",
    fill_value=0
)

# Sort dates
pivot_df = pivot_df.sort_index(axis=1)

# Format column names
pivot_df.columns = pivot_df.columns.strftime("%d-%b")

# Add totals
pivot_df["Total Hours"] = pivot_df.sum(axis=1)

# Reset index
pivot_df = pivot_df.reset_index()
timer.log_task("Create pivot table", time.time() - pivot_start)

# ----------------------------
# CREATE EXCEL
# ----------------------------
excel_start = time.time()
wb = Workbook()
ws = wb.active
ws.title = "Employee Hours"

# Write data
for r in dataframe_to_rows(pivot_df, index=False, header=True):
    ws.append(r)

# ----------------------------
# STYLING
# ----------------------------
header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")
    cell.fill = header_fill

# Freeze panes (lock names while scrolling dates)
ws.freeze_panes = "C2"

# Auto column width
for col in ws.columns:
    max_length = 0
    col_letter = col[0].column_letter
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    ws.column_dimensions[col_letter].width = max_length + 2

# ----------------------------
# SAVE FILE
# ----------------------------
output_file = os.path.join(output_folder, "Employee_Daily_Hours_Formatted.xlsx")
wb.save(output_file)
timer.log_task("Create and save Excel", time.time() - excel_start)

print("Saved:", output_file)
timer.save()